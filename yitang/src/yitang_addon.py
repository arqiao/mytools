"""逐字稿信息拓展工具 - 从字幕和讨论区中挖掘逐字稿遗漏的信息"""

import argparse
import json
import logging
import re
import time
from pathlib import Path

import requests
import yaml
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
CFG_DIR = PROJECT_DIR / "cfg"
LOG_DIR = PROJECT_DIR / "log-err"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "addon.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 配置加载
# ---------------------------------------------------------------------------

def load_config():
    """加载 addon 配置和凭证"""
    with open(CFG_DIR / "config-addon.yaml", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    with open(CFG_DIR / "credentials.yaml", encoding="utf-8") as f:
        creds = yaml.safe_load(f)
    return config, creds


# ---------------------------------------------------------------------------
# 文件解析
# ---------------------------------------------------------------------------

def parse_srt(filepath):
    """解析 SRT 字幕文件，返回 [(序号, 开始时间, 结束时间, 文本), ...]"""
    text = Path(filepath).read_text(encoding="utf-8")
    blocks = re.split(r"\n\s*\n", text.strip())
    entries = []
    for block in blocks:
        lines = block.strip().split("\n")
        if len(lines) < 3:
            continue
        seq = lines[0].strip()
        time_line = lines[1].strip()
        content = "\n".join(lines[2:]).strip()
        m = re.match(r"(\d[\d:,]+)\s*-->\s*(\d[\d:,]+)", time_line)
        if m:
            entries.append((int(seq), m.group(1), m.group(2), content))
    log.info(f"字幕解析完成: {len(entries)} 条, 来源: {filepath}")
    return entries


def srt_to_text(entries):
    """将 SRT 条目合并为纯文本（保留时间标记便于溯源）"""
    chunks = []
    for seq, start, end, text in entries:
        chunks.append(f"[{start}] {text}")
    return "\n".join(chunks)


def parse_discussion_xlsx(filepath):
    """解析讨论区 xlsx，返回 [{时间, 发言人, 内容, 标签...}, ...]"""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        item = {}
        for i, val in enumerate(row):
            if i < len(headers):
                item[headers[i]] = str(val) if val else ""
        # 跳过空内容
        content = item.get("发言内容", "").strip()
        if content:
            result.append(item)
    log.info(f"讨论区解析完成: {len(result)} 条有效发言, 来源: {filepath}")
    return result


def parse_transcript(filepath):
    """解析逐字稿文件（md/txt），按章节拆分。
    返回 [(章节标题, 章节内容), ...]"""
    text = Path(filepath).read_text(encoding="utf-8")
    # 按 markdown 标题拆分
    sections = []
    current_title = "(开头)"
    current_lines = []
    for line in text.split("\n"):
        m = re.match(r"^(#{1,6})\s+(.+)", line)
        if m:
            if current_lines:
                sections.append((current_title, "\n".join(current_lines)))
            current_title = m.group(2).strip()
            current_lines = [line]
        else:
            current_lines.append(line)
    if current_lines:
        sections.append((current_title, "\n".join(current_lines)))
    log.info(f"逐字稿解析完成: {len(sections)} 个章节, 来源: {filepath}")
    return sections


# ---------------------------------------------------------------------------
# LLM 调用层（OpenAI 兼容接口，支持 minimax/volcengine/dashscope/deepseek）
# ---------------------------------------------------------------------------

class LLMClient:
    """统一的 LLM 调用客户端，基于 OpenAI 兼容的 chat/completions 接口"""

    def __init__(self, config, creds):
        provider = config["llm"]["provider"]
        llm_cfg = config["llm"][provider]
        self.model = llm_cfg["model"]
        self.base_url = llm_cfg["base_url"].rstrip("/")
        self.max_tokens = llm_cfg.get("max_tokens", 8192)
        self.temperature = llm_cfg.get("temperature", 0.3)
        self.provider = provider

        # 获取 API key
        cred_block = creds.get(provider, {})
        self.api_key = cred_block.get("api_key", "")
        if not self.api_key:
            raise ValueError(f"未配置 {provider} 的 api_key，请检查 credentials.yaml")

        # 火山方舟用 endpoint_id 作为 model
        if provider == "volcengine":
            ep_id = cred_block.get("endpoint_id", "")
            if ep_id:
                self.model = ep_id

        # 统计
        self.total_calls = 0
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        log.info(f"LLM 客户端初始化: provider={provider}, model={self.model}")

    def chat(self, system_prompt, user_prompt, temperature=None):
        """发送 chat 请求，返回助手回复文本"""
        url = f"{self.base_url}/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}",
        }
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "max_tokens": self.max_tokens,
            "temperature": temperature or self.temperature,
        }

        for attempt in range(3):
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=300)
                resp.raise_for_status()
                data = resp.json()
                # 统计 token
                usage = data.get("usage", {})
                self.total_calls += 1
                self.total_input_tokens += usage.get("prompt_tokens", 0)
                self.total_output_tokens += usage.get("completion_tokens", 0)
                reply = data["choices"][0]["message"]["content"]
                return reply
            except requests.exceptions.HTTPError as e:
                if resp.status_code == 429:
                    wait = 2 ** attempt * 5
                    log.warning(f"速率限制，等待 {wait}s 后重试...")
                    time.sleep(wait)
                    continue
                log.error(f"LLM 调用失败: {e}, 响应: {resp.text}")
                raise
            except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectionError) as e:
                if attempt < 2:
                    wait = 2 ** attempt * 10
                    log.warning(f"LLM 调用超时/连接错误，等待 {wait}s 后重试 ({attempt+1}/3)...")
                    time.sleep(wait)
                    continue
                log.error(f"LLM 调用超时，重试耗尽: {e}")
                raise
            except Exception as e:
                log.error(f"LLM 调用异常: {e}")
                raise
        raise RuntimeError("LLM 调用重试次数耗尽")

    def report_usage(self):
        """输出 token 用量统计"""
        log.info(
            f"LLM 用量统计: provider={self.provider}, model={self.model}, "
            f"调用 {self.total_calls} 次, "
            f"输入 {self.total_input_tokens} tokens, "
            f"输出 {self.total_output_tokens} tokens"
        )


# ---------------------------------------------------------------------------
# 提示词
# ---------------------------------------------------------------------------

def load_prompt(filename):
    """从 cfg 目录加载提示词文件"""
    prompt_path = CFG_DIR / filename
    if not prompt_path.exists():
        raise FileNotFoundError(f"提示词文件不存在: {prompt_path}")
    return prompt_path.read_text(encoding="utf-8").strip()


# ---------------------------------------------------------------------------
# JSON 解析与结果处理
# ---------------------------------------------------------------------------

def parse_llm_json(reply):
    """从 LLM 回复中提取 JSON 数组。
    处理 markdown code fence 包裹的情况，解析失败返回空列表。"""
    text = reply.strip()
    # 尝试提取 code fence 中的 JSON
    m = re.search(r"```(?:json)?\s*\n(.*?)\n```", text, re.DOTALL)
    if m:
        text = m.group(1).strip()
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
        log.warning("LLM 返回的 JSON 不是数组，忽略")
        return []
    except json.JSONDecodeError as e:
        log.error(f"JSON 解析失败: {e}")
        # 保存原始回复到日志目录便于排查
        dump_path = LOG_DIR / f"llm_raw_{int(time.time())}.txt"
        dump_path.write_text(reply, encoding="utf-8")
        log.error(f"原始回复已保存: {dump_path}")
        return []


def merge_results(subtitle_items, discussion_items, sections):
    """按章节合并字幕和讨论区提取结果。
    sections 中可能有同名章节，用索引区分。
    返回 (merged_list, stats)：
    - merged_list: [(章节标题, {"subtitle": [...], "discussion": [...]}), ...]
      与 sections 一一对应
    - stats: 统计数据 dict
    """
    # 初始化：与 sections 一一对应
    merged = [{"subtitle": [], "discussion": []} for _ in sections]

    def add_item(item, source):
        chapter = item.get("chapter", "")
        idx = _match_chapter_idx(chapter, sections)
        merged[idx][source].append(item)

    for item in subtitle_items:
        add_item(item, "subtitle")
    for item in discussion_items:
        add_item(item, "discussion")

    # 去重：同一章节同一来源内，内容高度相似的条目只保留第一条
    dedup_count = 0
    for data in merged:
        for source in ("subtitle", "discussion"):
            items = data[source]
            if len(items) <= 1:
                continue
            unique = []
            for item in items:
                content = item.get("content", "")
                if not any(_is_similar(content, u.get("content", "")) for u in unique):
                    unique.append(item)
                else:
                    dedup_count += 1
            data[source] = unique
    if dedup_count:
        log.info(f"去重移除 {dedup_count} 条重复信息")

    # 统计
    sub_cats = {}
    disc_cats = {}
    for data in merged:
        for it in data["subtitle"]:
            cat = it.get("category", "其他")
            sub_cats[cat] = sub_cats.get(cat, 0) + 1
        for it in data["discussion"]:
            cat = it.get("category", "其他")
            disc_cats[cat] = disc_cats.get(cat, 0) + 1

    total_sub = sum(sub_cats.values())
    total_disc = sum(disc_cats.values())
    # 信息密度最高的章节（同名章节各自独立统计）
    density = []
    for i, (title, _) in enumerate(sections):
        n = len(merged[i]["subtitle"]) + len(merged[i]["discussion"])
        if n > 0:
            density.append((i, title, n))
    density.sort(key=lambda x: x[2], reverse=True)

    stats = {
        "total_subtitle": total_sub,
        "total_discussion": total_disc,
        "subtitle_categories": sub_cats,
        "discussion_categories": disc_cats,
        "top_chapters": [(t, n) for _, t, n in density[:5]],
    }
    return merged, stats


def _match_chapter_idx(chapter, sections):
    """模糊匹配章节，返回 sections 中的索引。
    精确匹配 > 包含匹配 > 归入第一个章节。"""
    if not chapter:
        return 0
    # 精确匹配
    for i, (title, _) in enumerate(sections):
        if chapter == title:
            return i
    # 包含匹配
    for i, (title, _) in enumerate(sections):
        if chapter in title or title in chapter:
            return i
    return 0


def _is_similar(a, b, threshold=0.6):
    """判断两段文本是否高度相似（基于字符集合的 Jaccard 相似度）。
    适用于检测 LLM 跨段重复输出的近似内容。"""
    if not a or not b:
        return False
    # 按字符 bigram 计算
    def bigrams(s):
        s = s.replace(" ", "")
        return set(s[i:i+2] for i in range(len(s) - 1)) if len(s) > 1 else {s}
    sa, sb = bigrams(a), bigrams(b)
    if not sa or not sb:
        return False
    intersection = len(sa & sb)
    union = len(sa | sb)
    return (intersection / union) >= threshold


def render_full_report(merged, stats, sections, prefix, output_path):
    """渲染完整版信息拓展 markdown 报告"""
    lines = []
    lines.append(f"# {prefix} — 信息拓展\n")

    # 摘要仪表盘
    lines.append("## 摘要统计\n")
    total = stats["total_subtitle"] + stats["total_discussion"]
    lines.append(f"- 提取信息总计：**{total}** 条")
    lines.append(f"  - 字幕补充：{stats['total_subtitle']} 条")
    lines.append(f"  - 讨论区精华：{stats['total_discussion']} 条")
    if stats["subtitle_categories"]:
        cats = "、".join(f"{k}({v})" for k, v in stats["subtitle_categories"].items())
        lines.append(f"  - 字幕分类：{cats}")
    if stats["discussion_categories"]:
        cats = "、".join(f"{k}({v})" for k, v in stats["discussion_categories"].items())
        lines.append(f"  - 讨论区分类：{cats}")
    if stats["top_chapters"]:
        top = "、".join(f"{t}({n}条)" for t, n in stats["top_chapters"][:3])
        lines.append(f"  - 信息密度最高：{top}")
    lines.append("")

    # 按章节输出
    for i, (title, _) in enumerate(sections):
        data = merged[i]
        sub_items = data["subtitle"]
        disc_items = data["discussion"]
        if not sub_items and not disc_items:
            continue

        lines.append(f"## {title}\n")

        if sub_items:
            lines.append("### 字幕补充\n")
            _render_subtitle_items(sub_items, lines)
            lines.append("")

        if disc_items:
            lines.append("### 讨论区精华\n")
            _render_discussion_items(disc_items, lines)
            lines.append("")

    content = "\n".join(lines)
    output_path.write_text(content, encoding="utf-8")
    log.info(f"完整版输出: {output_path}")
    return content


def _render_subtitle_items(items, lines):
    """渲染字幕补充条目"""
    # 按类别分组
    by_cat = {}
    for it in items:
        cat = it.get("category", "其他")
        by_cat.setdefault(cat, []).append(it)
    for cat, cat_items in by_cat.items():
        lines.append(f"**{cat}**\n")
        for it in cat_items:
            t = it.get("time", "")
            content = it.get("content", "")
            context = it.get("context", "")
            correction = it.get("correction", "")
            line = f"- [{t}] {content}"
            if context:
                line += f"\n  - 上下文：{context}"
            if correction:
                line += f"\n  - 纠正：{correction}"
            lines.append(line)
        lines.append("")


def _render_discussion_items(items, lines):
    """渲染讨论区精华条目"""
    by_cat = {}
    for it in items:
        cat = it.get("category", "其他")
        by_cat.setdefault(cat, []).append(it)
    for cat, cat_items in by_cat.items():
        lines.append(f"**{cat}**\n")
        for it in cat_items:
            t = it.get("time", "")
            speaker = it.get("speaker", "")
            tags = it.get("tags", "")
            content = it.get("content", "")
            tag_str = f"（{tags}）" if tags else ""
            line = f"- [{t}] **{speaker}**{tag_str}：{content}"
            lines.append(line)
            # 渲染对话串
            thread = it.get("thread", [])
            if thread:
                for msg in thread:
                    mt = msg.get("time", "")
                    ms = msg.get("speaker", "")
                    mc = msg.get("content", "")
                    lines.append(f"  - [{mt}] {ms}：{mc}")
        lines.append("")


# ---------------------------------------------------------------------------
# 分段与分析
# ---------------------------------------------------------------------------

def chunk_srt_text(srt_entries, chunk_size=30000):
    """将字幕条目按字符数分段，返回 [(起始时间, 结束时间, 文本), ...]"""
    chunks = []
    current_text = []
    current_len = 0
    start_time = srt_entries[0][1] if srt_entries else "00:00:00"
    end_time = start_time

    for seq, s_time, e_time, text in srt_entries:
        line = f"[{s_time}] {text}"
        if current_len + len(line) > chunk_size and current_text:
            chunks.append((start_time, end_time, "\n".join(current_text)))
            current_text = []
            current_len = 0
            start_time = s_time
        current_text.append(line)
        current_len += len(line) + 1
        end_time = e_time

    if current_text:
        chunks.append((start_time, end_time, "\n".join(current_text)))

    return chunks


def format_discussion_text(discussions):
    """将讨论区数据格式化为文本"""
    lines = []
    for d in discussions:
        time_str = d.get("时间", "")
        name = d.get("发言人名称", "")
        content = d.get("发言内容", "")
        tags = []
        for tag_key in ("MBA", "年级", "NPC等级", "进步阶梯", "其他标签"):
            v = d.get(tag_key, "").strip()
            if v:
                tags.append(v)
        tag_str = f" ({', '.join(tags)})" if tags else ""
        lines.append(f"[{time_str}] {name}{tag_str}: {content}")
    return "\n".join(lines)


def chunk_discussions(discussions, chunk_size=30000):
    """将讨论区发言按字符数分段，返回 [(段序号, 文本), ...]"""
    chunks = []
    current_lines = []
    current_len = 0
    for d in discussions:
        line = format_discussion_text([d])
        if current_len + len(line) > chunk_size and current_lines:
            chunks.append("\n".join(current_lines))
            current_lines = []
            current_len = 0
        current_lines.append(line)
        current_len += len(line) + 1
    if current_lines:
        chunks.append("\n".join(current_lines))
    return chunks


def is_feishu_url(path_or_url):
    """判断输入是飞书文档 URL 还是本地文件路径"""
    return "feishu.cn/" in path_or_url and path_or_url.startswith("http")


def fetch_feishu_transcript(url, creds):
    """通过飞书 API 获取文档内容，转为章节列表。
    复用 yitang_wiki.py 的 YitangCopier 能力。"""
    from yitang_wiki import YitangCopier
    copier = YitangCopier()

    # 判断是 wiki 链接还是 docx 链接
    if "/wiki/" in url:
        doc_id = copier.resolve_wiki_token(url)
        docx_url = f"https://arqiaoknow.feishu.cn/docx/{doc_id}"
        blocks_data = copier.fetch_feishu_blocks(docx_url)
    else:
        blocks_data = copier.fetch_feishu_blocks(url)

    # 展平 blocks
    root = blocks_data.get("blocks", {})
    flat = copier._flatten_blocks(root)
    log.info(f"飞书文档获取完成: {len(flat)} 个 blocks")

    # 用 _block_to_md 转为 markdown，再按章节拆分
    lines = []
    for b in flat:
        md = copier._block_to_md(b)
        if md:
            lines.append(md)
    full_text = "\n\n".join(lines)

    # 按 markdown 标题拆分为章节
    sections = []
    current_title = "(开头)"
    current_lines = []
    for line in full_text.split("\n"):
        m = re.match(r"^(#{1,6})\s+(.+)", line)
        if m:
            if current_lines:
                sections.append((current_title, "\n".join(current_lines)))
            current_title = m.group(2).strip()
            current_lines = [line]
        else:
            current_lines.append(line)
    if current_lines:
        sections.append((current_title, "\n".join(current_lines)))

    log.info(f"飞书逐字稿解析完成: {len(sections)} 个章节")
    return sections


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="逐字稿信息拓展工具")
    parser.add_argument("--dry-run", action="store_true",
                        help="试运行：只解析文件，不调用 LLM")
    parser.add_argument("--provider", type=str, default=None,
                        help="覆盖配置中的 LLM 提供商")
    parser.add_argument("--subtitle-only", action="store_true",
                        help="仅分析字幕（跳过讨论区）")
    parser.add_argument("--discussion-only", action="store_true",
                        help="仅分析讨论区（跳过字幕）")
    parser.add_argument("--no-digest", action="store_true",
                        help="跳过精华摘要生成")
    args = parser.parse_args()

    config, creds = load_config()

    # 命令行覆盖 provider
    if args.provider:
        config["llm"]["provider"] = args.provider

    # 验证输入文件
    inputs = config.get("input", {})
    transcript_src = inputs.get("transcript", "")
    subtitle_path = inputs.get("subtitle", "")
    discussion_path = inputs.get("discussion", "")

    # 相对路径基于 output.dir（即 localscript）目录
    base_dir = PROJECT_DIR / config.get("output", {}).get("dir", "localscript")

    def resolve_path(p):
        """相对路径拼接 base_dir，绝对路径直接用"""
        if not p or is_feishu_url(p):
            return p
        pp = Path(p)
        if pp.is_absolute():
            return str(pp)
        # 先试原路径，再试 base_dir 下
        if pp.exists():
            return str(pp)
        resolved = base_dir / pp
        return str(resolved)

    subtitle_path = resolve_path(subtitle_path)
    discussion_path = resolve_path(discussion_path)

    # 优先使用 _fix 版本的字幕（经过 srtfix 校订的版本）
    if subtitle_path and not is_feishu_url(subtitle_path):
        sp = Path(subtitle_path)
        fix_path = sp.parent / f"{sp.stem}_fix{sp.suffix}"
        if fix_path.exists():
            log.info(f"检测到校订版字幕，优先使用: {fix_path.name}")
            subtitle_path = str(fix_path)

    if not transcript_src:
        log.error("未配置逐字稿来源 (input.transcript)")
        return

    # 解析逐字稿：支持飞书 URL 或本地文件
    if is_feishu_url(transcript_src):
        log.info(f"逐字稿来源: 飞书文档 {transcript_src}")
        _, creds_data = load_config()
        sections = fetch_feishu_transcript(transcript_src, creds_data)
    elif Path(transcript_src).exists():
        log.info(f"逐字稿来源: 本地文件 {transcript_src}")
        sections = parse_transcript(transcript_src)
    else:
        log.error(f"逐字稿文件不存在: {transcript_src}")
        return

    section_titles = [t for t, _ in sections]
    log.info(f"章节列表: {section_titles}")

    # 解析字幕
    srt_entries = []
    if subtitle_path and Path(subtitle_path).exists():
        srt_entries = parse_srt(subtitle_path)
        log.info(f"字幕文本总长: {sum(len(e[3]) for e in srt_entries)} 字符")
    elif not args.discussion_only:
        log.warning(f"字幕文件不存在或未配置: {subtitle_path}")

    # 解析讨论区
    discussions = []
    if discussion_path and Path(discussion_path).exists():
        discussions = parse_discussion_xlsx(discussion_path)
    elif not args.subtitle_only:
        log.warning(f"讨论区文件不存在或未配置: {discussion_path}")

    if args.dry_run:
        log.info("=== 试运行模式，不调用 LLM ===")
        log.info(f"逐字稿: {len(sections)} 章节")
        log.info(f"字幕: {len(srt_entries)} 条")
        log.info(f"讨论区: {len(discussions)} 条发言")
        return

    # 初始化 LLM 客户端
    llm = LLMClient(config, creds)
    analysis_cfg = config.get("analysis", {})
    prompt_subtitle = load_prompt(analysis_cfg.get("prompt_subtitle", "prompt-subtitle.md"))
    prompt_discussion = load_prompt(analysis_cfg.get("prompt_discussion", "prompt-discussion.md"))
    chunk_size = analysis_cfg.get("chunk_size", 30000)
    output_dir = Path(config.get("output", {}).get("dir", "localscript"))
    if not output_dir.is_absolute():
        output_dir = PROJECT_DIR / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    # 生成输出文件名前缀：优先字幕文件名 > 讨论区文件名 > 逐字稿
    prefix = config.get("output", {}).get("prefix", "")
    if not prefix:
        if subtitle_path and Path(subtitle_path).exists():
            # 去掉 _wm 后缀，如 AI落地Live_069_wm.srt → AI落地Live_069
            stem = Path(subtitle_path).stem
            prefix = re.sub(r"_wm$", "", stem)
        elif discussion_path and Path(discussion_path).exists():
            prefix = Path(discussion_path).stem
        elif not is_feishu_url(transcript_src):
            prefix = Path(transcript_src).stem
        else:
            prefix = "feishu_doc"

    # ── 字幕对比分析 ──
    subtitle_items = []
    if srt_entries and not args.discussion_only:
        log.info("=== 开始字幕对比分析 ===")
        transcript_text = "\n\n".join(
            f"## {title}\n{content}" for title, content in sections
        )
        chunks = chunk_srt_text(srt_entries, chunk_size)
        log.info(f"字幕分为 {len(chunks)} 段")

        for i, (start, end, chunk_text) in enumerate(chunks):
            log.info(f"  分析字幕段 {i+1}/{len(chunks)}: {start} ~ {end}")
            user_prompt = (
                f"## 逐字稿内容\n\n{transcript_text}\n\n"
                f"---\n\n"
                f"## 字幕原文（{start} ~ {end}）\n\n{chunk_text}"
            )
            try:
                reply = llm.chat(prompt_subtitle, user_prompt)
                items = parse_llm_json(reply)
                subtitle_items.extend(items)
                log.info(f"    提取 {len(items)} 条信息")
            except Exception as e:
                log.error(f"    字幕段 {i+1} 处理失败，跳过: {e}")

        log.info(f"字幕对比分析完成，共提取 {len(subtitle_items)} 条")

    # ── 讨论区分析 ──
    discussion_items = []
    if discussions and not args.subtitle_only:
        log.info("=== 开始讨论区分析 ===")
        # 章节标题 + 首段摘要，提升关联准确度
        section_info = []
        for title, content in sections:
            # 取前 200 字作为摘要
            plain = re.sub(r"^#{1,6}\s+.*\n?", "", content).strip()
            summary = plain[:200] + ("..." if len(plain) > 200 else "")
            section_info.append(f"- {title}：{summary}")
        section_list = "\n".join(section_info)

        disc_chunks = chunk_discussions(discussions, chunk_size)
        log.info(f"讨论区分为 {len(disc_chunks)} 段")

        for i, chunk_text in enumerate(disc_chunks):
            log.info(f"  分析讨论区段 {i+1}/{len(disc_chunks)}")
            user_prompt = (
                f"## 逐字稿章节（标题 + 摘要）\n\n{section_list}\n\n"
                f"---\n\n"
                f"## 讨论区发言记录（第 {i+1}/{len(disc_chunks)} 段）\n\n{chunk_text}"
            )
            try:
                reply = llm.chat(prompt_discussion, user_prompt)
                items = parse_llm_json(reply)
                discussion_items.extend(items)
                log.info(f"    提取 {len(items)} 条信息")
            except Exception as e:
                log.error(f"    讨论区段 {i+1} 处理失败，跳过: {e}")

        log.info(f"讨论区分析完成，共提取 {len(discussion_items)} 条")

    # ── 合并与渲染 ──
    if subtitle_items or discussion_items:
        merged, stats = merge_results(subtitle_items, discussion_items, sections)
        report_path = output_dir / f"{prefix}_信息拓展.md"
        report_content = render_full_report(
            merged, stats, sections, prefix, report_path
        )

        # ── 精华摘要 ──
        if not args.no_digest:
            log.info("=== 生成精华摘要 ===")
            prompt_digest = load_prompt(
                analysis_cfg.get("prompt_digest", "prompt-digest.md")
            )
            digest_reply = llm.chat(prompt_digest, report_content)
            # 清理 LLM 可能包裹的 code fence
            digest_text = re.sub(
                r"^```(?:markdown)?\s*\n(.*)\n```\s*$", r"\1",
                digest_reply.strip(), flags=re.DOTALL,
            )
            digest_path = output_dir / f"{prefix}_精华摘要.md"
            digest_path.write_text(digest_text, encoding="utf-8")
            log.info(f"精华摘要输出: {digest_path}")

    llm.report_usage()
    log.info("完成")


if __name__ == "__main__":
    main()
