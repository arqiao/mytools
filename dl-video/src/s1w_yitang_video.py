"""直播视频下载及讨论区导出工具"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import hashlib
import hmac
import json
import logging
import re
import time
from base64 import b64decode, b64encode
from pathlib import Path

import requests
import yaml
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives.padding import PKCS7

from modules.ffmpeg_utils import extract_audio
from modules.config_utils import load_config

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
CFG_DIR = PROJECT_DIR / "cfg"
_input_cfg = yaml.safe_load((CFG_DIR / "input.yaml").read_text(encoding="utf-8")) or {}
LOG_DIR = PROJECT_DIR / _input_cfg.get("path_log_dir", "log-err")
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "video_download.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# 签名常量（与 yitang_copier.py 相同）
G_KEY = b"BDFHJLNPRTVXZ\\^`"
M_KEY = b"ether7sv6te7sv6he7sv6there7sv6r0"

# 标签 ID 到名称的映射（从 API /api/air/room/icon-list 获取）
ICON_MAP = {
    # 特殊标签
    2: "学习委员",
    3: "我请客",
    5: "出圈er",
    10: "TA",
    21: "1班学习委员",
    22: "2班学习委员",
    23: "3班学习委员",
    2202: "top",  # 进步阶梯
    2203: "MBA",  # MBA 标识
    2422: "加速十年",
    # SNPC
    9001: "SNPC-1-1",
    9002: "SNPC-1-2",
    9003: "SNPC-1-3",
    9004: "SNPC-1-4",
    9011: "SNPC-2-1",
    9012: "SNPC-2-2",
    9013: "SNPC-2-3",
    9014: "SNPC-2-4",
    9021: "SNPC-3-1",
    9022: "SNPC-3-2",
    9023: "SNPC-3-3",
    9024: "SNPC-3-4",
    # 进步阶梯 (grade)
    10050: "进步阶梯-50",
    10090: "进步阶梯-90",
    10200: "进步阶梯-200",
    10500: "进步阶梯-500",
    10900: "进步阶梯-900",
    11783: "进步阶梯-1783",
    # MBA (study)
    40002: "MBA",
    40003: "MBA-提前批",
    40004: "MBA-正常批",
    40005: "MBA-5期",
    40006: "MBA-6期",
    40007: "MBA-7期",
    40008: "MBA-8期",
    40009: "年级-25级",
    40010: "年级-26级",
    # 线下报名
    50001: "报名北京线下",
    50002: "报名青岛线下",
    50003: "报名太原线下",
    50004: "报名天津线下",
    50005: "报名上海线下",
    50006: "报名杭州线下",
    50007: "报名南京线下",
    50008: "报名深圳线下",
    50009: "报名广州线下",
    50010: "报名厦门线下",
    50011: "报名福州线下",
    50012: "报名武汉线下",
    50013: "报名长沙线下",
    50014: "报名郑州线下",
    50015: "报名西安线下",
    50016: "报名成都线下",
    50017: "报名重庆线下",
    50018: "报名贵阳线下",
    50019: "报名昆明线下",
    50020: "报名合肥线下",
    50021: "报名沈阳线下",
}


class YitangLiveDownloader:
    def __init__(self):
        config, creds = load_config()
        self.config = config
        self.creds = creds
        self.session = requests.Session()
        self.icon_map = dict(ICON_MAP)  # 复制基础映射
        self._load_icon_map()  # 从 API 加载动态映射
        self.output_prefix = self.config["s1_yitang_ailive"]["output_prefix"]
        self.series_name = self.config["s1_yitang_ailive"]["query_copystr"]
        self.output_dir = PROJECT_DIR / self.config["path_yitang_dir"]

    def _load_icon_map(self):
        """从 API 获取 icon 列表，仅补充静态映射中没有的条目"""
        try:
            uri = "/api/air/room/icon-list"
            params = {"roomId": ""}
            url = f"https://air.yitang.top{uri}"
            headers = self._get_headers(uri, params)
            resp = self.session.get(url, params=params, headers=headers, timeout=15)
            data = resp.json()
            if data.get("code") == 0:
                for item in data.get("data", {}).get("list", []):
                    icon_id = item.get("id")
                    name = item.get("name", "")
                    if icon_id and name and icon_id not in self.icon_map:
                        self.icon_map[icon_id] = name
                log.info(f"已加载 {len(self.icon_map)} 个 icon 标签")
        except Exception as e:
            log.warning(f"加载 icon 列表失败，使用默认映射: {e}")

    # ── 签名生成 ──────────────────────────────────────────

    def _aes_encrypt(self, plaintext: bytes, key: bytes, iv: bytes) -> str:
        cipher = Cipher(algorithms.AES(key), modes.CBC(iv))
        encryptor = cipher.encryptor()
        padder = PKCS7(128).padder()
        padded = padder.update(plaintext) + padder.finalize()
        ct = encryptor.update(padded) + encryptor.finalize()
        return b64encode(ct).decode()

    def _generate_x_token_1(self) -> str:
        token = self.creds["yitang"]["token"]
        ts = str(int(time.time()))
        plaintext = f"{G_KEY.decode()}~{token}~{ts}".encode()
        key = hashlib.md5(G_KEY).hexdigest().encode()
        return self._aes_encrypt(plaintext, key, G_KEY)

    def _generate_x_token_2(self, uri: str, params: dict) -> str:
        sorted_qs = "&".join(f"{k}={params[k]}" for k in sorted(params))
        message = f"{uri}?{sorted_qs}"
        sig = hmac.new(M_KEY, message.encode(), hashlib.sha1).hexdigest()
        return sig

    # ── API 调用 ──────────────────────────────────────────

    def _get_headers(self, uri: str, params: dict) -> dict:
        return {
            "Authorization": f"Bearer {self.creds['yitang']['token']}",
            "Cookie": self.creds["yitang"]["cookie"],
            "X-token-1": self._generate_x_token_1(),
            "X-token-2": self._generate_x_token_2(uri, params),
            "X-Ctxid": self.creds["yitang"]["request_id"],
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }

    def fetch_replay_data(self, live_id: str) -> dict:
        """获取直播回放数据"""
        uri = "/api/air/room/replay"
        params = {
            "sid": live_id,
            "lang": "zh-CN",
            "_uds": "onair_yitang_index_code4",
        }
        url = f"https://air.yitang.top{uri}"
        headers = self._get_headers(uri, params)
        resp = self.session.get(url, params=params, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") != 0:
            raise RuntimeError(f"API 错误: {data}")
        return data.get("data", {})

    def fetch_live_title(self, source_url: str, live_id: str) -> str:
        """获取直播标题：优先 room/info API，其次 lesson/detail API"""

        # 1. room/info API（用 sid 参数，最可靠）
        try:
            uri = "/api/air/room/info"
            params = {"sid": live_id}
            headers = self._get_headers(uri, params)
            resp = self.session.get(
                f"https://air.yitang.top{uri}", params=params, headers=headers, timeout=15
            )
            data = json.loads(resp.content.decode("utf-8"))
            if data.get("code") == 0:
                title = data["data"].get("title", "")
                if title:
                    return title
        except Exception as e:
            log.warning(f"room/info 获取标题失败: {e}")

        # 2. lesson/detail API（URL 带 lessonId 时）
        lesson_id_match = re.search(r"lessonId=([^&]+)", source_url)
        if lesson_id_match:
            lesson_id = lesson_id_match.group(1)
            try:
                uri = "/api/lesson/detail"
                params = {"lessonId": lesson_id}
                headers = self._get_headers(uri, params)
                resp = self.session.get(
                    f"https://air.yitang.top{uri}", params=params, headers=headers, timeout=15
                )
                data = json.loads(resp.content.decode("utf-8"))
                if data.get("code") == 0:
                    name = data.get("data", {}).get("name", "")
                    if name:
                        return name
            except Exception as e:
                log.warning(f"lesson/detail 获取标题失败: {e}")

        return "一堂直播"

    def extract_number_from_title(self, title: str) -> str:
        """从标题中提取数字编号，如 'AI落地Live第69场' -> '069'"""
        # 匹配 Live 后面的数字（支持 Live069、Live第69场 等格式）
        match = re.search(r'Live[第]?(\d+)', title)
        if match:
            return match.group(1).zfill(3)
        return ""

    # ── 视频下载 ──────────────────────────────────────────

    def download_video(self, video_url: str, output_path: Path) -> bool:
        """下载视频（m3u8 -> ts），yt-dlp 自带断点续传"""
        try:
            import subprocess
            output_ts = output_path.with_suffix(".ts")
            if output_ts.exists():
                log.info(f"视频已存在，跳过: {output_ts}")
                return True
            cmd = [
                "yt-dlp",
                "-o", str(output_ts),
                video_url,
            ]
            result = subprocess.run(cmd)
            if result.returncode == 0 and output_ts.exists():
                log.info(f"视频下载成功: {output_ts}")
                return True
            log.warning("视频下载失败")
            return False
        except FileNotFoundError:
            log.warning("未安装 yt-dlp")
            return False

    def _download_m3u8(self, m3u8_url: str, output_path: Path) -> bool:
        """直接下载 m3u8（仅下载音频或分段视频）"""
        try:
            resp = self.session.get(m3u8_url, timeout=30)
            resp.raise_for_status()
            content = resp.text

            # 如果是 m3u8 播放列表，尝试获取音频文件
            if "#EXTM3U" in content:
                # 简单处理：获取第一个音频/视频片段
                for line in content.split("\n"):
                    if line and not line.startswith("#"):
                        # 如果是相对路径，拼接完整 URL
                        if not line.startswith("http"):
                            base = m3u8_url.rsplit("/", 1)[0]
                            line = f"{base}/{line}"
                        # 下载第一个片段作为测试
                        log.info(f"检测到 m3u8 播放列表，实际内容需要 yt-dlp 处理")
                        return False

            # 保存为 .m3u8 文件
            m3u8_path = output_path.with_suffix(".m3u8")
            m3u8_path.write_text(content, encoding="utf-8")
            log.warning(f"已保存 m3u8 文件: {m3u8_path}，请使用 yt-dlp 下载")
            return False
        except Exception as e:
            log.error(f"下载失败: {e}")
            return False

    def download_audio(self, audio_url: str, output_path: Path) -> bool:
        """下载音频（mp3），支持断点续传"""
        try:
            mp3_path = output_path.with_suffix(".mp3")
            req_headers = {}
            downloaded = 0
            mode = "wb"

            if mp3_path.exists():
                downloaded = mp3_path.stat().st_size
                head = self.session.head(audio_url, timeout=15)
                total_size = int(head.headers.get("content-length", 0))
                if total_size and downloaded >= total_size:
                    log.info(f"音频已存在，跳过: {mp3_path}")
                    return mp3_path
                req_headers["Range"] = f"bytes={downloaded}-"
                mode = "ab"
                log.info(f"音频续传，已下载 {downloaded / 1024 / 1024:.1f} MB")

            resp = self.session.get(audio_url, stream=True, headers=req_headers, timeout=300)
            resp.raise_for_status()
            total = int(resp.headers.get("content-length", 0)) + downloaded
            with open(mp3_path, mode) as f:
                last_reported = downloaded // (5 * 1024 * 1024)
                for chunk in resp.iter_content(chunk_size=1024 * 256):
                    f.write(chunk)
                    downloaded += len(chunk)
                    cur_block = downloaded // (5 * 1024 * 1024)
                    if total and cur_block > last_reported:
                        last_reported = cur_block
                        pct = downloaded * 100 // total
                        print(f"\r  音频下载: {pct}% ({downloaded / 1024 / 1024:.1f} / "
                              f"{total / 1024 / 1024:.1f} MB)", end="", flush=True)
            if total:
                print(f"\r  音频下载: 100% ({total / 1024 / 1024:.1f} / "
                      f"{total / 1024 / 1024:.1f} MB)")
            log.info(f"音频下载成功: {mp3_path}")
            return mp3_path
        except Exception as e:
            log.error(f"音频下载失败: {e}")
            return None

    # ── 讨论区导出 ────────────────────────────────────────

    def export_chats(self, chats: list, output_path: Path, duration: int):
        """导出讨论区内容到 CSV"""
        if not chats:
            log.warning("讨论区无内容")
            return

        # 将 offset（秒）转换为时间字符串（统一 HH:MM:SS 格式）
        def offset_to_time(offset_seconds: int) -> str:
            hours = offset_seconds // 3600
            minutes = (offset_seconds % 3600) // 60
            seconds = offset_seconds % 60
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

        # 解析 icons 为标签字符串
        def parse_icons(icons: list) -> str:
            if not icons:
                return ""
            tags = []
            for icon_id in icons:
                tag = self.icon_map.get(icon_id, f"icon-{icon_id}")
                tags.append(tag)
            return "； ".join(tags)

        # 分离 MBA、年级、NPC等级、进步阶梯、其他标签
        def parse_user_tags(icons: list) -> dict:
            has_mba_flag = False  # 2203 MBA 身份标识
            mba_list = []
            grade_list = []
            npc_list = []
            progress_list = []
            other_list = []

            for icon_id in icons:
                if icon_id == 2203:
                    has_mba_flag = True
                    continue
                name = self.icon_map.get(icon_id, f"icon-{icon_id}")
                if "MBA" in name:
                    mba_list.append(name)
                # 年级: 年级 + grade-xxx
                elif "年级" in name or name.startswith("grade-"):
                    grade_list.append(name)
                # NPC等级: NPC 或 SNPC 开头的
                elif "NPC" in name or name.startswith("SNPC"):
                    npc_list.append(name)
                # 进步阶梯
                elif "进步阶梯" in name:
                    progress_list.append(name)
                # 其他标签：学习委员、我请客、线下报名等
                else:
                    other_list.append(name)

            # 只有 2203 没有具体期数时，补一个 "MBA"
            if has_mba_flag and not mba_list:
                mba_list.append("MBA")

            return {
                "MBA": "； ".join(mba_list),
                "年级": "； ".join(grade_list),
                "NPC等级": "； ".join(npc_list),
                "进步阶梯": "； ".join(progress_list),
                "其他标签": "； ".join(other_list),
            }

        def _clean_cell_text(text):
            """过滤 emoji 和 XML 非法控制字符"""
            if not text:
                return ""
            return "".join(
                c for c in text
                if ord(c) < 0x10000
                and (ord(c) >= 0x20 or c in "\t\n\r")
            )

        rows = []
        for chat in chats:
            tags_info = parse_user_tags(chat.get("icons", []))

            rows.append({
                "时间": offset_to_time(chat.get("offset", 0)),
                "发言人名称": _clean_cell_text(chat.get("name", "")),
                "发言内容": _clean_cell_text(chat.get("message", "")),
                "MBA": tags_info["MBA"],
                "年级": tags_info["年级"],
                "NPC等级": tags_info["NPC等级"],
                "进步阶梯": tags_info["进步阶梯"],
                "其他标签": tags_info["其他标签"],
            })

        # 按时间排序
        rows.sort(key=lambda x: x["时间"])

        # 写入 XLSX
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "讨论区"

            # 写入表头
            headers = ["时间", "发言人名称", "发言内容", "MBA", "年级", "NPC等级", "进步阶梯", "其他标签"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # 写入数据
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

            # 设置列宽
            ws.column_dimensions["A"].width = 10  # 时间
            ws.column_dimensions["B"].width = 15  # 发言人名称
            ws.column_dimensions["C"].width = 40  # 发言内容
            ws.column_dimensions["D"].width = 12  # MBA
            ws.column_dimensions["E"].width = 12  # 年级
            ws.column_dimensions["F"].width = 12  # NPC等级
            ws.column_dimensions["G"].width = 15  # 进步阶梯
            ws.column_dimensions["H"].width = 20  # 其他标签

            wb.save(output_path)
            log.info(f"讨论区已导出: {output_path} ({len(rows)} 条记录)")
        except PermissionError:
            log.warning(f"XLSX 文件被占用，跳过导出: {output_path}")

    # ── 主流程 ────────────────────────────────────────────

    def run(self, livestreams=None, dl_video=True, dl_audio=True, dl_chat=True):
        if livestreams is None:
            # 独立运行：从 tasks 中筛选 yitang 任务
            all_tasks = self.config.get("tasks", [])
            livestreams = [{"source_url": t["source_huifang_url"],
                            "title": t.get("title", ""),
                            "output_name": t.get("output_name", "")}
                           for t in all_tasks
                           if t.get("source_type") == "yitang" and t.get("source_huifang_url")]
        if not livestreams:
            log.warning("未找到 yitang 视频任务")
            return

        output_dir = self.output_dir
        output_dir.mkdir(exist_ok=True)

        for idx, item in enumerate(livestreams):
            source_url = item["source_url"]

            log.info(f"[{idx + 1}/{len(livestreams)}] 处理: {source_url}")

            # 从 URL 提取 liveId（去掉查询参数）
            live_id_match = re.search(r'/live/([^/?]+)', source_url)
            if not live_id_match:
                log.warning(f"  无法从URL提取liveId: {source_url}")
                continue
            live_id = live_id_match.group(1)
            log.info(f"  Live ID: {live_id}")

            try:
                # 1. 获取回放数据
                data = self.fetch_replay_data(live_id)
                replay = data.get("replay", {})
                chats = data.get("chats", [])
                duration = replay.get("duration", 0)

                if not replay.get("url"):
                    log.warning("  无回放视频，跳过")
                    continue

                # 获取直播标题：优先使用配置文件中手动指定的标题
                manual_title = item.get("title", "").strip()
                if manual_title:
                    title = manual_title
                else:
                    title = self.fetch_live_title(source_url, live_id)
                log.info(f"  标题: {title}")

                # 确定输出文件名：优先使用配置中手动指定的 output_name
                manual_output = item.get("output_name", "").strip()
                if manual_output:
                    output_name = manual_output
                elif self.series_name in title:
                    # 使用前缀+编号
                    title_number = self.extract_number_from_title(title)
                    if not title_number:
                        title_number = str(idx + 1).zfill(3)
                    output_name = f"{self.output_prefix}{title_number}"
                else:
                    # 使用标题作为文件名（只替换 Windows 不允许的半角字符）
                    safe_title = re.sub(r'[\\/:*?"<>|：]', '_', title)
                    safe_title = safe_title.strip()
                    if not safe_title:
                        safe_title = f"直播_{live_id}"
                    output_name = safe_title

                video_url = replay["url"]
                audio_url = replay.get("audioUrl", "")

                log.info(f"  视频: {video_url}")
                log.info(f"  讨论区: {len(chats)} 条")

                video_path = output_dir / output_name

                # 2. 导出讨论区（轻量操作优先）
                if dl_chat and chats:
                    xlsx_path = output_dir / f"{output_name}.xlsx"
                    if xlsx_path.exists():
                        log.info(f"  讨论区已存在，跳过: {xlsx_path}")
                    else:
                        self.export_chats(chats, xlsx_path, duration)

                # 3. 下载视频
                if dl_video:
                    self.download_video(video_url, video_path)

                # 4. 下载音频（依赖视频文件，放在视频之后）
                if dl_audio:
                    got_audio = False
                    if audio_url:
                        got_audio = self.download_audio(audio_url, video_path)
                    if not got_audio:
                        log.info("服务端无独立音频文件，从本地视频提取")
                        video_file = video_path.with_suffix(".ts")
                        mp3_path = video_path.with_suffix(".mp3")
                        if video_file.exists():
                            extract_audio(video_file, mp3_path)
                        else:
                            log.warning("视频文件不存在，无法提取音频")

            except Exception as e:
                log.error(f"  处理失败: {e}", exc_info=True)
                continue

        log.info("全部处理完成")


def process_yitang_video(task, config, creds):
    """供 s1_huifang 调度器调用的入口"""
    downloader = YitangLiveDownloader()
    livestream = {"source_url": task["source_huifang_url"],
                  "title": task.get("title", ""),
                  "output_name": task.get("output_name", "")}
    downloader.run(livestreams=[livestream], dl_video=True, dl_audio=True, dl_chat=True)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="直播视频下载及讨论区导出")
    parser.add_argument("--video", action="store_true", help="只下载视频")
    parser.add_argument("--audio", action="store_true", help="只下载音频")
    parser.add_argument("--chat", action="store_true", help="只导出讨论区")
    args = parser.parse_args()

    # 都不带则全部生成
    if not args.video and not args.audio and not args.chat:
        args.video = args.audio = args.chat = True

    downloader = YitangLiveDownloader()
    downloader.run(dl_video=args.video, dl_audio=args.audio, dl_chat=args.chat)
