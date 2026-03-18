#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
SAT 词汇表数据提取器（示例版）
在 Windows / macOS / Linux 下均可运行。
"""

import pandas as pd


def parse_vocabulary_content():
    """
    示例：返回一小段 SAT 词汇数据，用于生成 Excel。
    真正的项目里可以改为从 PDF 中解析得到完整数据。
    """
    sample_vocabulary = [
        {
            "英文单词": "a heap of",
            "中文释义": "phrase. 一堆",
            "章节单元": "list-01",
            "音标": "",
            "记忆标记": "Ⓜ①②③④⑤⑥⑦",
        },
        {
            "英文单词": "accidental",
            "中文释义": "adj. 偶然的",
            "章节单元": "list-01",
            "音标": "/ˌæksɪˈdentl/",
            "记忆标记": "Ⓜ①②③④⑤⑥⑦",
        },
        {
            "英文单词": "acknowledgement",
            "中文释义": "n. 承认；感谢",
            "章节单元": "list-01",
            "音标": "/əkˈnɒlɪdʒmənt/",
            "记忆标记": "Ⓜ①②③④⑤⑥⑦",
        },
        {
            "英文单词": "imperishable",
            "中文释义": "adj. 不会腐烂的；不坏的；不朽的",
            "章节单元": "list-02",
            "音标": "/ɪmˈperɪʃəbl/",
            "记忆标记": "Ⓜ①②③④⑤⑥⑦",
        },
        {
            "英文单词": "take for granted",
            "中文释义": "phrase. 认为……是理所应当",
            "章节单元": "list-03",
            "音标": "",
            "记忆标记": "Ⓜ①②③④⑤⑥⑦",
        },
    ]

    return sample_vocabulary


def create_sat_vocabulary_excel(filename: str = "SAT_词汇表_示例.xlsx") -> str:
    """
    创建 SAT 词汇表 Excel 文件（带简单样式，如果安装了 openpyxl）。
    """
    print("📊 正在创建 SAT 词汇表 Excel 文件……")

    # 获取词汇数据
    vocabulary_data = parse_vocabulary_content()

    # 创建 DataFrame
    df = pd.DataFrame(vocabulary_data)

    # 重新排列列顺序
    df = df[["英文单词", "中文释义", "章节单元", "音标", "记忆标记"]]

    try:
        # 尝试使用 openpyxl 加样式
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="SAT 词汇表", index=False)
            worksheet = writer.sheets["SAT 词汇表"]

            # 设置列宽
            worksheet.column_dimensions["A"].width = 20  # 英文单词
            worksheet.column_dimensions["B"].width = 45  # 中文释义
            worksheet.column_dimensions["C"].width = 15  # 章节单元
            worksheet.column_dimensions["D"].width = 25  # 音标
            worksheet.column_dimensions["E"].width = 20  # 记忆标记

            # 标题行样式
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # 数据行左对齐
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # 边框
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # 行高
            for row_idx in range(1, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 25

        print(f"✅ Excel 文件创建成功：{filename}")
        print(f"📈 示例词汇数量：{len(df)}")
        print(f"📚 包含章节：{sorted(df['章节单元'].unique())}")
        return filename

    except ImportError:
        # 如果没有安装 openpyxl，则退化为基础版本
        fallback_name = "SAT_词汇表_基础.xlsx"
        df.to_excel(fallback_name, index=False)
        print(f"✅ Excel 文件创建成功（基础版本）：{fallback_name}")
        return fallback_name


def generate_full_data_structure() -> None:
    """
    在控制台输出数据结构和使用说明。
    """
    print("\n" + "=" * 60)
    print("📋 SAT 词汇表数据结构说明")
    print("=" * 60)

    print("\n📊 数据统计：")
    print("- 总章节数：25 个（list-01 到 list-25）")
    print("- 预计总词汇数：3500–4000 个（真实项目中由 PDF 实际解析得到）")
    print("- 词汇类型：单词、短语")
    print("- 记忆标记：Ⓜ①②③④⑤⑥⑦ 系统")

    print("\n📋 表格列结构：")
    print("1. 英文单词（20 列宽）")
    print("   - 单一单词：'accidental'")
    print("   - 短语：'a heap of', 'take for granted'")

    print("\n2. 中文释义（45 列宽）")
    print("   - 格式：'词性. 中文含义'")
    print("   - 示例：'adj. 偶然的', 'n. 承认；感谢'")
    print("   - 短语标记：'phrase. 一堆'")

    print("\n3. 章节单元（15 列宽）")
    print("   - 格式：'list-01' 到 'list-25'")
    print("   - 对应原书：第 7 页开始到第 154 页")

    print("\n4. 音标（25 列宽）")
    print("   - 国际音标格式：'/ˌæksɪˈdentl/'")
    print("   - 短语无音标：空值")

    print("\n5. 记忆标记（20 列宽）")
    print("   - 格式：'Ⓜ①②③④⑤⑥⑦'")
    print("   - 含义：永久记忆到第七轮记忆")

    print("\n🎯 使用方法：")
    print("1. 打开生成的 Excel 文件")
    print("2. 按章节学习：从 list-01 开始")
    print("3. 标记已掌握：隐藏或删除已学会的词汇")
    print("4. 进度追踪：定期检查掌握情况")

    print("\n📝 建议扩展：")
    print("- 添加例句列")
    print("- 添加记忆方法列")
    print("- 添加掌握程度列（1–5 分）")
    print("- 添加复习日期列")


if __name__ == "__main__":
    print("🚀 SAT 词汇表 Excel 生成器")
    print("=" * 50)

    # 创建 Excel 文件
    output_filename = create_sat_vocabulary_excel()

    # 显示数据结构说明
    generate_full_data_structure()

    print(f"\n✅ 处理完成！文件已保存为：{output_filename}")
    print("📥 请打开 Excel 文件进行学习使用。")

